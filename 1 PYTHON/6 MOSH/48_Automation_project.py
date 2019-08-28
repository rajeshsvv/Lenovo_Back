# Autmation related to spread sheet to update 100 or 10000 of excel spreadsheets at a time.


'''
import openpyxl as xl
from openpyxl.chart import BarChart,Reference


wb=xl.load_workbook("transactions.xlsx")
sheet=wb["Sheet1"]
cell=sheet['a1']
# print(cell.value)
# print(sheet.max_row)     #4
# print(sheet.cell(1,1))

for row in range(2,sheet.max_row+1):
    # print(row)
    cell=sheet.cell(row,3)
    # print(cell.value)
    corrected_price=cell.value*10
    corrected_price_cell=sheet.cell(row,4)
    corrected_price_cell.value=corrected_price

values=Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4
          )

chart=BarChart()
chart.add_data(values)
# sheet.add_chart(chart,'e2')       # u got key error if u give e for co-ordinate in excel sheet.
sheet.add_chart(chart,'E2')

wb.save('transactions2.xlsx')
'''

# organised code

def process_workbook(filename):
    import openpyxl as xl
    from openpyxl.chart import BarChart, Reference

    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    cell = sheet['a1']


    for row in range(2, sheet.max_row + 1):
        # print(row)
        cell = sheet.cell(row, 3)
        # print(cell.value)
        corrected_price = cell.value * 10
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4
                       )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    wb.save('transactions2.xlsx')